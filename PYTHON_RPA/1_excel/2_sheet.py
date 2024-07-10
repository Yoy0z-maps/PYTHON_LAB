from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 활성화된 sheet 뒤에 새로운 sheet를 기본 이름으로 생성
ws.title = 'MySheet'
ws.sheet_properties.tabColor = 'ff66ff' #RGB로 sheet 탭 색상 변경

ws1 = wb.create_sheet('YourSheet') # 주어진 인수를 이름으로 새로운 Sheet 생성
ws2 = wb.create_sheet('IndexSheet', 0) # 0번째 인덱스에 새로운 시트 생성

# sheet에 접근하는 법
# ws1.method
''' dictionary 활용
name_ws = wb['sheetName']
name_ws.method
'''

# Sheet 복사
new_ws = wb["IndexSheet"]
new_ws['A1'] = 'Test' # A1 셀 값에 Test 입력
target = wb.copy_worksheet(new_ws) # variable target에 new_ws를 카피해서 새로운 시트 생성
target.title = 'Copied Sheet' # 새로 생성한 시트 이름을 변경


print(wb.sheetnames) # 모든 sheet 이름 확인

wb.save('sample.xlsx')