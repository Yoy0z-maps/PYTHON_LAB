from openpyxl import Workbook
from random import *

from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기 (가로)
ws.append(['번호','영어','수학'])
for i in range(1,11):
    ws.append([i,randint(0,100),randint(0,100)])

# 특정 데이터만 가져오기
#Column
col_B = ws['B'] # B(영어) column만 가져오기
print(col_B)
for cell in col_B:
    print(cell.value)

# 여러개의 데이터 동시에 가져오기
col_range = ws['B:C'] # B(영어) C (수학) column 가져오기
for cols in col_range:
    for cell in cols:
        print(cell.value)

#Row
row_stu1 = ws[2]
for cell in row_stu1:
    print(cell.value)

row_range = ws[1:6] #1에서 6번줄까지 가져오기 여기서는 끝번호가 -1된 값이 아니다!!!
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

# 마지막 데이터를 모를 때
row_range2 = ws[2:ws.max_row] # 마지막까지 (column의 경우 max_col)

for rows in row_range2:
    for cell in rows:
        print(cell.coordinate, end=' ') # 셀에 대한 좌표 정보를 출력한다. coordinate from import openpyxl.utility.cell
        xy = coordinate_from_string(cell.coordinate) # 각 cell의 좌표마다 tuple형태로 row, column 좌표를 나눠준다
        print(xy, end='')
        print(xy[0], end=" ") # row alphabet
        print(xy[1], end=" ") # column number
    print()

# 전체 rows (행)
print(tuple(ws.rows))

# 전체 columns (열)
print(tuple(ws.columns))

for row in tuple(ws.rows):
    print(row[1].value) # 각 Row의 2번째 cell만 출력

for column in tuple(ws.columns):
    print(column[0].value) # 각 column의 1번째 cell 출력

# 전체 rows 2
for row in ws.iter_rows():
    print(row[0].value)
# 전체 column 2
for column in ws.iter_cols():
    print(column[0].value)

#58:27 2번째행부터 11번째행까지, 2번째열부터 3번째열까지
for row in ws.iter_rows(min_row=2,max_row=11,min_col=2,max_col=3):
    print(row[0].value, row[1].value)

wb.save('sample.xlsx')