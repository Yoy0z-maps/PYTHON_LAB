from openpyxl import Workbook;
wb = Workbook()
ws = wb.active
ws.title = 'NadoSheet'

#workSheetName['Alphabet(Row)Number(Column)'] => 셀에 값 입력하기
ws['A1'] = 1
ws['B2'] = 2
ws['C3'] = 3

ws.cell(column=4,row=4,value=4)

# 셀의 값 읽어오기
print(ws['A1']) # 해당 셀의 객체 정보를 출력
print(ws['A2'])

print(ws['B2'].value) # 해당 셀의 값만 출력
print(ws['B1'].value)

'''
row = A, B, C, D...
column = 1, 2, 3, 4...
'''
print(ws.cell(row=3,column=3).value) # print(ws['C3'].value)와 같다


# for문과 같이 써보기
from random import *
for x in range(11,21): # 10개 row
    for y in range(11,21): # 10개 column
        ws.cell(row=x,column=y,value=randint(0,100))

index = 1

for x in range(31,41):
    for y in range(31,41):
        ws.cell(row=x,column=y,value=index)
        index += 1


wb.save('sample.xlsx')